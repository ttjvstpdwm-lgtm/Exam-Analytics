from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
import re
from typing import BinaryIO

import numpy as np
import pandas as pd
from openpyxl import load_workbook


QUESTION_ID_RE = re.compile(r"^Question ID\s+(\d+)$", re.IGNORECASE)
EXTRA_COLUMN_PREFIX = "excel_col__"

STUDENT_COLUMNS = [
    "Username",
    "Last Name",
    "First Name",
    "Full Name",
    "CLASSE",
]

FIXED_STUDENT_COLUMNS = {
    "username",
    "last name",
    "first name",
    "full name",
    "classe",
    "totale chiuse",
    "totale aperte",
    "progetto",
    "totale esame",
    "finale non arrotondato",
    "finale rounded",
    "visione compiti",
    "visione",
    "iscritto visione",
    "iscritta visione",
    "iscrizione visione",
    "prenotato visione",
    "prenotata visione",
    "prenotazione visione",
}


@dataclass(frozen=True)
class ExamData:
    source_name: str
    exam_sheets: list[str]
    long_df: pd.DataFrame
    student_df: pd.DataFrame


def normalize_label(value: object) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def coerce_number(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def excel_source_from_bytes(file_bytes: bytes) -> BytesIO:
    return BytesIO(file_bytes)


def _series_or_blank(df: pd.DataFrame, column: str) -> pd.Series:
    resolved = _resolve_column(df, column)
    if resolved is not None:
        return df[resolved]
    return pd.Series([np.nan] * len(df), index=df.index)


def _series_or_blank_any(df: pd.DataFrame, *columns: str) -> pd.Series:
    for column in columns:
        resolved = _resolve_column(df, column)
        if resolved is not None:
            return df[resolved]
    return pd.Series([np.nan] * len(df), index=df.index)


def _resolve_column(df: pd.DataFrame, column: str) -> str | None:
    normalized = str(column).strip().lower()
    for candidate in df.columns:
        if str(candidate).strip().lower() == normalized:
            return candidate
    return None


def _question_numbers(columns: pd.Index) -> list[int]:
    nums: list[int] = []
    for column in columns:
        match = QUESTION_ID_RE.match(str(column))
        if match:
            nums.append(int(match.group(1)))
    return sorted(set(nums))


def _is_question_detail_column(column: object) -> bool:
    text = str(column).strip()
    if QUESTION_ID_RE.match(text):
        return True
    return bool(
        re.match(
            r"^(Question|Answer|Possible Points|Auto Score|Manual Score|Correct Answer|Risposta corretta|Soluzione)\s+\d+$",
            text,
            re.IGNORECASE,
        )
        or re.match(r"^Commento\s*Q?\s*\d+$", text, re.IGNORECASE)
    )


def _extra_column_key(column: object, used_keys: set[str]) -> str:
    base = EXTRA_COLUMN_PREFIX + str(column).strip()
    key = base
    counter = 2
    while key in used_keys:
        key = f"{base} ({counter})"
        counter += 1
    used_keys.add(key)
    return key


def _extra_student_columns(df: pd.DataFrame) -> pd.DataFrame:
    extras: dict[str, pd.Series] = {}
    used_keys: set[str] = set()
    for column in df.columns:
        normalized = str(column).strip().lower()
        if normalized in FIXED_STUDENT_COLUMNS or _is_question_detail_column(column):
            continue
        extras[_extra_column_key(column, used_keys)] = df[column]
    if not extras:
        return pd.DataFrame(index=df.index)
    return pd.DataFrame(extras, index=df.index)


def find_exam_sheets(source: str | Path | BytesIO | BinaryIO) -> list[str]:
    xl = pd.ExcelFile(source)
    return _matching_exam_sheets(xl)


def _matching_exam_sheets(xl: pd.ExcelFile, sheet_names: list[str] | tuple[str, ...] | None = None) -> list[str]:
    candidate_sheets = xl.sheet_names if sheet_names is None else list(dict.fromkeys(sheet_names))
    sheets: list[str] = []
    for sheet in candidate_sheets:
        if sheet not in xl.sheet_names:
            continue
        header = pd.read_excel(xl, sheet_name=sheet, nrows=0)
        if _question_numbers(header.columns):
            sheets.append(sheet)
    return sheets


def load_exam(
    source: str | Path | BytesIO | BinaryIO,
    source_name: str = "Esame",
    sheet_names: list[str] | tuple[str, ...] | None = None,
) -> ExamData:
    xl = pd.ExcelFile(source)
    requested_sheets = list(dict.fromkeys(sheet_names)) if sheet_names else None
    if requested_sheets:
        missing_sheets = [sheet for sheet in requested_sheets if sheet not in xl.sheet_names]
        if missing_sheets:
            raise ValueError("I fogli selezionati non esistono nel file: " + ", ".join(missing_sheets))
    exam_sheets = _matching_exam_sheets(xl, requested_sheets)
    long_parts: list[pd.DataFrame] = []
    student_parts: list[pd.DataFrame] = []

    for sheet in exam_sheets:
        df = pd.read_excel(xl, sheet_name=sheet)
        username_all = _series_or_blank(df, "Username").map(normalize_label)
        full_name_all = _series_or_blank(df, "Full Name").map(normalize_label)
        has_student = username_all.ne("") | full_name_all.ne("")
        df = df.loc[has_student].copy()
        question_numbers = _question_numbers(df.columns)
        source_row = pd.Series(df.index, index=df.index, name="source_row")
        username = _series_or_blank(df, "Username").map(normalize_label)
        student_uid = sheet + "|" + source_row.astype(str) + "|" + username.astype(str)
        extra_student_columns = _extra_student_columns(df)

        student_part = pd.DataFrame(
            {
                "student_uid": student_uid,
                "sheet_name": sheet,
                "source_row": source_row,
                "username": username,
                "last_name": _series_or_blank(df, "Last Name").map(normalize_label),
                "first_name": _series_or_blank(df, "First Name").map(normalize_label),
                "full_name": _series_or_blank(df, "Full Name").map(normalize_label),
                "classe": _series_or_blank(df, "CLASSE").map(normalize_label),
                "blackboard_total_closed": coerce_number(_series_or_blank(df, "TOTALE CHIUSE")),
                "totale_aperte": coerce_number(_series_or_blank(df, "TOTALE APERTE")),
                "progetto": coerce_number(_series_or_blank(df, "PROGETTO")),
                "totale_esame": coerce_number(_series_or_blank(df, "TOTALE ESAME")),
                "finale_non_arrotondato": coerce_number(_series_or_blank(df, "FINALE NON ARROTONDATO")),
                "finale_rounded": coerce_number(_series_or_blank(df, "FINALE ROUNDED")),
                "visione_compiti": _series_or_blank_any(
                    df,
                    "VISIONE COMPITI",
                    "VISIONE",
                    "ISCRITTO VISIONE",
                    "ISCRITTA VISIONE",
                    "ISCRIZIONE VISIONE",
                    "PRENOTATO VISIONE",
                    "PRENOTATA VISIONE",
                    "PRENOTAZIONE VISIONE",
                ).map(normalize_label),
            }
        )
        if not extra_student_columns.empty:
            student_part = pd.concat([student_part, extra_student_columns], axis=1)
        student_parts.append(student_part)

        for question_num in question_numbers:
            possible = coerce_number(_series_or_blank(df, f"Possible Points {question_num}"))
            auto_score = coerce_number(_series_or_blank(df, f"Auto Score {question_num}"))
            manual_score = coerce_number(_series_or_blank(df, f"Manual Score {question_num}"))
            part = pd.DataFrame(
                {
                    "student_uid": student_uid,
                    "sheet_name": sheet,
                    "source_row": source_row,
                    "username": username,
                    "last_name": _series_or_blank(df, "Last Name").map(normalize_label),
                    "first_name": _series_or_blank(df, "First Name").map(normalize_label),
                    "full_name": _series_or_blank(df, "Full Name").map(normalize_label),
                    "classe": _series_or_blank(df, "CLASSE").map(normalize_label),
                    "question_num": question_num,
                    "question_id": _series_or_blank(df, f"Question ID {question_num}").map(normalize_label),
                    "question_text": _series_or_blank(df, f"Question {question_num}").map(normalize_label),
                    "answer": _series_or_blank(df, f"Answer {question_num}").map(normalize_label),
                    "correct_answer": _series_or_blank_any(
                        df,
                        f"Correct Answer {question_num}",
                        f"Correct answer {question_num}",
                        f"Risposta corretta {question_num}",
                        f"Risposta Corretta {question_num}",
                        f"Soluzione {question_num}",
                    ).map(normalize_label),
                    "comment": _series_or_blank_any(
                        df,
                        f"Commento Q{question_num}",
                        f"Commento{question_num}",
                        f"Commento {question_num}",
                    ).map(normalize_label),
                    "possible_points": possible,
                    "auto_score": auto_score,
                    "manual_score_file": manual_score,
                    "blackboard_total_closed": coerce_number(_series_or_blank(df, "TOTALE CHIUSE")),
                    "totale_aperte": coerce_number(_series_or_blank(df, "TOTALE APERTE")),
                    "progetto": coerce_number(_series_or_blank(df, "PROGETTO")),
                    "totale_esame": coerce_number(_series_or_blank(df, "TOTALE ESAME")),
                    "finale_non_arrotondato": coerce_number(_series_or_blank(df, "FINALE NON ARROTONDATO")),
                    "finale_rounded": coerce_number(_series_or_blank(df, "FINALE ROUNDED")),
                    "visione_compiti": _series_or_blank_any(
                        df,
                        "VISIONE COMPITI",
                        "VISIONE",
                        "ISCRITTO VISIONE",
                        "ISCRITTA VISIONE",
                        "ISCRIZIONE VISIONE",
                        "PRENOTATO VISIONE",
                        "PRENOTATA VISIONE",
                        "PRENOTAZIONE VISIONE",
                    ).map(normalize_label),
                }
            )
            if not extra_student_columns.empty:
                part = pd.concat([part, extra_student_columns], axis=1)
            long_parts.append(part)

    if requested_sheets and not long_parts:
        raise ValueError("Nessuno dei fogli selezionati contiene colonne 'Question ID n'.")
    if not long_parts:
        raise ValueError("Non ho trovato fogli con colonne 'Question ID n'.")

    long_df = pd.concat(long_parts, ignore_index=True)
    student_df = pd.concat(student_parts, ignore_index=True).drop_duplicates("student_uid")
    return ExamData(source_name=source_name, exam_sheets=exam_sheets, long_df=long_df, student_df=student_df)


def load_corrections(path: Path) -> pd.DataFrame:
    columns = [
        "student_uid",
        "sheet_name",
        "username",
        "full_name",
        "classe",
        "question_num",
        "manual_score",
        "updated_at",
    ]
    if not path.exists():
        return pd.DataFrame(columns=columns)
    df = pd.read_csv(path, dtype={"student_uid": str, "question_num": int})
    for column in columns:
        if column not in df.columns:
            df[column] = np.nan
    df["manual_score"] = pd.to_numeric(df["manual_score"], errors="coerce")
    df["question_num"] = pd.to_numeric(df["question_num"], errors="coerce").astype("Int64")
    return df[columns].dropna(subset=["student_uid", "question_num"])


def apply_corrections(long_df: pd.DataFrame, corrections: pd.DataFrame) -> pd.DataFrame:
    df = long_df.copy()
    df["question_num"] = df["question_num"].astype(int)
    if corrections.empty:
        df["manual_score"] = df["manual_score_file"]
    else:
        corr = corrections[["student_uid", "question_num", "manual_score"]].copy()
        corr["question_num"] = corr["question_num"].astype(int)
        corr = corr.dropna(subset=["manual_score"])
        corr = corr.drop_duplicates(["student_uid", "question_num"], keep="last")
        df = df.merge(
            corr.rename(columns={"manual_score": "manual_score_saved"}),
            on=["student_uid", "question_num"],
            how="left",
        )
        df["manual_score"] = df["manual_score_saved"].combine_first(df["manual_score_file"])
        df = df.drop(columns=["manual_score_saved"])
    return add_score_columns(df)


def add_score_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["auto_score"] = coerce_number(out["auto_score"])
    out["manual_score"] = coerce_number(out["manual_score"])
    out["auto_score_filled"] = out["auto_score"].fillna(0)
    out["integrated_score_raw"] = out["manual_score"].where(out["manual_score"].notna(), out["auto_score"])
    out["integrated_score"] = out["integrated_score_raw"].fillna(0)
    out["is_missing_score"] = out["integrated_score_raw"].isna()
    out["score_pct"] = np.where(
        out["possible_points"].fillna(0) > 0,
        out["integrated_score"] / out["possible_points"],
        np.nan,
    )

    conditions = [
        out["is_missing_score"],
        out["possible_points"].isna(),
        (out["possible_points"] > 0) & (out["integrated_score"] >= out["possible_points"]),
        out["integrated_score"].fillna(0).eq(0),
    ]
    choices = ["Da correggere", "Da verificare", "Corretta", "Errata"]
    out["status"] = np.select(conditions, choices, default="Parziale")
    return out


def student_totals(scored_df: pd.DataFrame) -> pd.DataFrame:
    grouped = scored_df.groupby(
        ["student_uid", "sheet_name", "source_row", "username", "last_name", "first_name", "full_name", "classe"],
        dropna=False,
    )
    totals = grouped.agg(
        totale_autoscore=("auto_score_filled", "sum"),
        totale_integrato=("integrated_score", "sum"),
        blackboard_total_closed=("blackboard_total_closed", "first"),
        totale_aperte=("totale_aperte", "first"),
        progetto=("progetto", "first"),
        totale_esame=("totale_esame", "first"),
        finale_non_arrotondato=("finale_non_arrotondato", "first"),
        finale_rounded=("finale_rounded", "first"),
        visione_compiti=("visione_compiti", "first"),
        punti_possibili=("possible_points", "sum"),
        domande_da_correggere=("is_missing_score", "sum"),
        domande_parziali=("status", lambda s: int((s == "Parziale").sum())),
        domande_errate=("status", lambda s: int((s == "Errata").sum())),
    ).reset_index()
    extra_cols = [column for column in scored_df.columns if str(column).startswith(EXTRA_COLUMN_PREFIX)]
    if extra_cols:
        extra_totals = grouped[extra_cols].first().reset_index()
        totals = totals.merge(
            extra_totals,
            on=["student_uid", "sheet_name", "source_row", "username", "last_name", "first_name", "full_name", "classe"],
            how="left",
        )
    totals["pct_integrato"] = np.where(
        totals["punti_possibili"].fillna(0) > 0,
        totals["totale_integrato"] / totals["punti_possibili"],
        np.nan,
    )
    totals["student_label"] = totals.apply(student_label, axis=1)
    return totals.sort_values(["classe", "last_name", "first_name", "username"], kind="stable")


def student_label(row: pd.Series) -> str:
    name = normalize_label(row.get("full_name"))
    username = normalize_label(row.get("username"))
    sheet = normalize_label(row.get("sheet_name"))
    classe = normalize_label(row.get("classe"))
    pieces = []
    if name:
        pieces.append(name)
    if username:
        pieces.append(username)
    suffix = " / ".join([p for p in [f"classe {classe}" if classe else "", sheet] if p])
    if suffix:
        pieces.append(f"({suffix})")
    return " ".join(pieces) if pieces else normalize_label(row.get("student_uid"))


def question_summary(scored_df: pd.DataFrame) -> pd.DataFrame:
    grouped = scored_df.groupby(["question_num", "question_id"], dropna=False)
    summary = grouped.agg(
        testo=("question_text", "first"),
        punti_possibili=("possible_points", "max"),
        media_autoscore=("auto_score", "mean"),
        media_integrata=("integrated_score", "mean"),
        studenti=("student_uid", "nunique"),
        da_correggere=("is_missing_score", "sum"),
        corrette=("status", lambda s: int((s == "Corretta").sum())),
        parziali=("status", lambda s: int((s == "Parziale").sum())),
        errate=("status", lambda s: int((s == "Errata").sum())),
    ).reset_index()
    summary["score_pct_medio"] = np.where(
        summary["punti_possibili"].fillna(0) > 0,
        summary["media_integrata"] / summary["punti_possibili"],
        np.nan,
    )
    summary["corrette_pct"] = np.where(summary["studenti"] > 0, summary["corrette"] / summary["studenti"], np.nan)
    return summary.sort_values("question_num")


def filter_by_class(scored_df: pd.DataFrame, selected_class: str) -> pd.DataFrame:
    if selected_class == "Tutte":
        return scored_df
    return scored_df[scored_df["classe"].astype(str) == str(selected_class)]


def save_student_corrections(path: Path, student_df: pd.DataFrame, edited_scores: pd.DataFrame) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    existing = load_corrections(path)
    student_uid = student_df["student_uid"].iloc[0]
    remaining = existing[existing["student_uid"] != student_uid].copy()

    rows: list[dict[str, object]] = []
    now = datetime.now().isoformat(timespec="seconds")
    for _, row in edited_scores.iterrows():
        score = row.get("manual_score")
        if pd.isna(score):
            continue
        rows.append(
            {
                "student_uid": student_uid,
                "sheet_name": student_df["sheet_name"].iloc[0],
                "username": student_df["username"].iloc[0],
                "full_name": student_df["full_name"].iloc[0],
                "classe": student_df["classe"].iloc[0],
                "question_num": int(row["question_num"]),
                "manual_score": float(score),
                "updated_at": now,
            }
        )

    new_rows = pd.DataFrame(rows, columns=existing.columns)
    if remaining.empty:
        updated = new_rows
    elif new_rows.empty:
        updated = remaining
    else:
        updated = pd.concat([remaining, new_rows], ignore_index=True)
    updated.to_csv(path, index=False)


def build_corrected_workbook(
    source: str | Path | bytes,
    corrections: pd.DataFrame,
    exam_sheets: list[str],
) -> bytes:
    if isinstance(source, bytes):
        workbook = load_workbook(BytesIO(source))
    else:
        workbook = load_workbook(source)

    corr = corrections.copy()
    if not corr.empty:
        corr["question_num"] = corr["question_num"].astype(int)
        corr = corr.dropna(subset=["manual_score"])

    for sheet_name in exam_sheets:
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        headers = {str(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if cell.value is not None}
        question_numbers = sorted(
            int(match.group(1))
            for header in headers
            for match in [QUESTION_ID_RE.match(header)]
            if match
        )

        total_auto_col = _ensure_header(ws, headers, "TOTALE AUTOSCORE APP")
        total_integrated_col = _ensure_header(ws, headers, "TOTALE INTEGRATO APP")
        missing_col = _ensure_header(ws, headers, "DOMANDE DA CORREGGERE APP")

        sheet_corr = corr[corr["sheet_name"] == sheet_name] if not corr.empty else corr
        correction_map = {
            (str(row["student_uid"]), int(row["question_num"])): row["manual_score"]
            for _, row in sheet_corr.iterrows()
        }

        username_col = headers.get("Username")
        for excel_row in range(2, ws.max_row + 1):
            source_row = excel_row - 2
            username = normalize_label(ws.cell(excel_row, username_col).value if username_col else "")
            student_uid = f"{sheet_name}|{source_row}|{username}"
            total_auto = 0.0
            total_integrated = 0.0
            missing = 0

            for question_num in question_numbers:
                auto_col = headers.get(f"Auto Score {question_num}")
                manual_col = headers.get(f"Manual Score {question_num}")
                auto_value = _num(ws.cell(excel_row, auto_col).value) if auto_col else np.nan

                if manual_col and (student_uid, question_num) in correction_map:
                    ws.cell(excel_row, manual_col).value = float(correction_map[(student_uid, question_num)])

                manual_value = _num(ws.cell(excel_row, manual_col).value) if manual_col else np.nan
                if not pd.isna(auto_value):
                    total_auto += float(auto_value)
                if not pd.isna(manual_value):
                    total_integrated += float(manual_value)
                elif not pd.isna(auto_value):
                    total_integrated += float(auto_value)
                else:
                    missing += 1

            ws.cell(excel_row, total_auto_col).value = round(total_auto, 4)
            ws.cell(excel_row, total_integrated_col).value = round(total_integrated, 4)
            ws.cell(excel_row, missing_col).value = missing

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _ensure_header(ws, headers: dict[str, int], header: str) -> int:
    if header in headers:
        return headers[header]
    col = ws.max_column + 1
    ws.cell(1, col).value = header
    headers[header] = col
    return col


def _num(value: object) -> float:
    try:
        if value is None or value == "":
            return np.nan
        return float(value)
    except (TypeError, ValueError):
        return np.nan
